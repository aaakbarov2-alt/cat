from datetime import timedelta
from io import BytesIO
from tempfile import TemporaryDirectory
from unittest.mock import patch

from django.contrib.auth.models import User
from django.db import IntegrityError, transaction
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from django.template import Context, Template

from .models import ExamSet, Question, QuestionGroup, Section, StudentAnswer, StudentAttempt
from .html_importer import create_exam_from_payload, parse_ielts_html
from .excel_importer import build_excel_template, parse_excel_test
from .views import SUBMISSION_GRACE_SECONDS
from .grading import academic_reading_band


IMPORT_HTML = b'''<!doctype html><html><body>
<div class="passage-content"><h2>Music history</h2><p>This is the reading passage.</p></div>
<div class="question-group"><p>Choose the correct answer.</p>
<div class="question" data-question="1"><div class="question-text">1. Which option is correct?</div><label><input type="radio" name="q1" value="A"> A. First choice</label><label><input type="radio" name="q1" value="B"> B. Second choice</label></div>
<div class="question" data-question="2"><div class="question-text">2. Complete this ____ <input type="text" name="q2">.</div></div>
</div><script>const correctAnswers = {q1: "B", q2: "answer"};</script></body></html>'''


class HtmlImportTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username="admin", email="admin@example.com", password="admin-pass-123"
        )

    def test_parser_extracts_real_question_types_and_answers(self):
        section = parse_ielts_html(IMPORT_HTML, "reading.html")

        self.assertEqual(section.passage_text, "Music history\n\nThis is the reading passage.")
        self.assertEqual(len(section.questions), 2)
        self.assertEqual(section.questions[0].question_type, "mcq")
        self.assertEqual(section.questions[0].options, ["A. First choice", "B. Second choice"])
        self.assertEqual(section.questions[0].correct_answer, "B. Second choice")
        self.assertEqual(section.questions[1].question_type, "gap")
        self.assertEqual(section.questions[1].correct_answer, "answer")

    def test_listening_import_does_not_require_a_reading_passage(self):
        listening_html = b'''<div class="question" data-question="1"><div class="question-text">1. Choose one.</div><label><input type="radio" value="A"> A</label><label><input type="radio" value="B"> B</label></div><script>const correctAnswers={q1:"B"};</script>'''
        section = parse_ielts_html(listening_html, "listening.html", section_type="listening")
        exam = create_exam_from_payload({"title":"Listening import", "description":"", "category":"listening", "section_type":"listening", "time_limit_minutes":10, "publish":False, "sections":[section.as_payload()]})

        self.assertEqual(exam.sections.get().section_type, "listening")
        self.assertEqual(exam.sections.get().questions.get().correct_answer, "B")

    def test_payload_creates_a_publishable_exam(self):
        section = parse_ielts_html(IMPORT_HTML, "reading.html")
        exam = create_exam_from_payload(
            {
                "title": "Imported reading test",
                "description": "Imported from HTML",
                "category": "reading",
                "section_type": "reading",
                "time_limit_minutes": 20,
                "publish": True,
                "sections": [section.as_payload()],
            }
        )

        self.assertTrue(exam.is_published)
        self.assertEqual(exam.sections.count(), 1)
        self.assertEqual(exam.sections.get().questions.count(), 2)

    def test_admin_home_uses_clean_unfold_dashboard(self):
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse("admin:index"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "IELTS Mock")
        self.assertContains(response, "Tests")
        self.assertContains(response, "/static/unfold/css/styles.css")

    def test_category_quick_create_prefills_the_exam_category(self):
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse("admin:exams_examset_add") + "?category=listening")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<option value="listening" selected>', html=False)


class ExactHtmlExamTests(TestCase):
    def setUp(self):
        self.student = User.objects.create_user(username="exact-student", password="test-pass-123")
        self.exam = ExamSet.objects.create(
            title="Original HTML test",
            category="reading",
            is_published=True,
            delivery_mode="exact_html",
            source_html="<!doctype html><html><body><h1 id='kept'>Keep me exact</h1></body></html>",
        )
        self.client.force_login(self.student)

    def test_exact_test_opens_in_the_dedicated_viewer_and_serves_raw_html(self):
        start = self.client.post(reverse("start_exam", args=[self.exam.id]))
        attempt = StudentAttempt.objects.get(student=self.student, exam_set=self.exam)
        self.assertRedirects(start, reverse("exact_exam", args=[attempt.id]), fetch_redirect_response=False)

        content = self.client.get(reverse("exact_exam_content", args=[attempt.id]))
        self.assertEqual(content.status_code, 200)
        self.assertEqual(content.content.decode(), self.exam.source_html)
        self.assertIn("sandbox", content.headers["Content-Security-Policy"])

    def test_exact_test_can_be_marked_complete(self):
        self.client.post(reverse("start_exam", args=[self.exam.id]))
        attempt = StudentAttempt.objects.get(student=self.student, exam_set=self.exam)

        response = self.client.post(reverse("complete_exact_exam", args=[attempt.id]))

        attempt.refresh_from_db()
        self.assertTrue(attempt.is_complete)
        self.assertRedirects(response, reverse("exam_list"), fetch_redirect_response=False)


class ReadingBandTests(TestCase):
    def test_full_reading_uses_the_40_question_band_table(self):
        self.assertEqual(academic_reading_band(40, 40), (9.0, 40, False))
        self.assertEqual(academic_reading_band(35, 40), (8.0, 35, False))
        self.assertEqual(academic_reading_band(30, 40), (7.0, 30, False))
        self.assertEqual(academic_reading_band(23, 40), (6.0, 23, False))

    def test_single_passage_keeps_raw_total_and_estimates_band(self):
        self.assertEqual(academic_reading_band(13, 13), (9.0, 40, True))
        self.assertEqual(academic_reading_band(10, 13), (7.0, 31, True))
        self.assertEqual(academic_reading_band(0, 13), (0.0, 0, True))


class ExcelImportTests(TestCase):
    def setUp(self):
        self.admin_user = User.objects.create_superuser(
            username="excel-admin", email="excel@example.com", password="admin-pass-123"
        )
        self.client.force_login(self.admin_user)

    def test_template_round_trip_creates_editable_native_test(self):
        workbook = build_excel_template()
        payload = parse_excel_test(workbook, publish=True)

        exam = create_exam_from_payload(payload)

        self.assertEqual(exam.delivery_mode, "native")
        self.assertEqual(exam.category, "reading")
        self.assertTrue(exam.is_published)
        self.assertEqual(exam.sections.count(), 1)
        self.assertEqual(exam.sections.get().questions.count(), 2)
        self.assertEqual(exam.sections.get().questions.get(order=1).explanation, "Explain why Option B is correct.")
        self.assertEqual(exam.sections.get().question_groups.count(), 1)
        self.assertEqual(exam.sections.get().questions.filter(group__key="notes_1").count(), 2)

    def test_group_renderer_replaces_excel_placeholders_with_question_inputs(self):
        payload = parse_excel_test(build_excel_template(), publish=True)
        exam = create_exam_from_payload(payload)
        section = exam.sections.get()
        group = section.question_groups.get(key="notes_1")
        rendered = Template("{% load exam_content %}{{ group|render_question_group }}").render(Context({"group": group}))
        self.assertIn('class="exam-question-group"', rendered)
        self.assertIn(f'name="q{section.questions.get(order=1).id}"', rendered)
        self.assertIn(f'id="question-{section.questions.get(order=2).id}"', rendered)
        self.assertNotIn("[[1]]", rendered)

    def test_excel_time_accepts_friendly_text_and_blank_defaults(self):
        from openpyxl import load_workbook
        workbook_file = build_excel_template()
        workbook = load_workbook(workbook_file)
        workbook["Sections"]["C2"] = "20 minutes"
        friendly = BytesIO()
        workbook.save(friendly)
        friendly.seek(0)
        self.assertEqual(parse_excel_test(friendly)["sections"][0]["time_limit_minutes"], 20)

        workbook["Sections"]["C2"] = ""
        blank = BytesIO()
        workbook.save(blank)
        blank.seek(0)
        self.assertEqual(parse_excel_test(blank)["sections"][0]["time_limit_minutes"], 60)

    def test_legacy_workbook_without_groups_still_imports(self):
        from openpyxl import load_workbook
        workbook = load_workbook(build_excel_template())
        del workbook["Groups"]
        workbook["Questions"].delete_cols(10)
        legacy = BytesIO()
        workbook.save(legacy)
        legacy.seek(0)
        payload = parse_excel_test(legacy)
        self.assertNotIn("groups", payload["sections"][0])
        self.assertTrue(all(not question.get("group_key") for question in payload["sections"][0]["questions"]))

    def test_preserved_excel_importer_creates_a_draft_exam(self):
        workbook = build_excel_template()
        payload = parse_excel_test(workbook, publish=False)
        exam = create_exam_from_payload(payload)

        self.assertEqual(exam.title, "Academic Reading Practice 1")
        self.assertFalse(exam.is_published)

    def test_plain_excel_content_renders_as_paragraphs_not_literal_html_tags(self):
        rendered = Template(
            "{% load exam_content %}<div>{{ value|render_rich_text }}</div>"
        ).render(Context({"value": "First paragraph.\n\nSecond paragraph."}))

        self.assertIn("<p>First paragraph.</p>", rendered)
        self.assertIn("<p>Second paragraph.</p>", rendered)
        self.assertNotIn("&lt;p&gt;", rendered)

class ExamConstraintTests(TestCase):
    def setUp(self):
        self.student = User.objects.create_user(username="student", password="test-pass-123")
        self.exam = ExamSet.objects.create(title="IELTS Mock 1")
        self.section = Section.objects.create(
            exam_set=self.exam,
            order=1,
            section_type="reading",
            time_limit_minutes=60,
            passage_text="The evidence sentence is in this reading passage.",
        )
        self.question = Question.objects.create(
            section=self.section,
            order=1,
            question_type="gap",
            prompt="Complete the sentence.",
            correct_answer="answer",
        )

    def test_only_one_active_attempt_is_allowed_per_student_and_exam(self):
        StudentAttempt.objects.create(student=self.student, exam_set=self.exam)

        with self.assertRaises(IntegrityError), transaction.atomic():
            StudentAttempt.objects.create(student=self.student, exam_set=self.exam)

    def test_completed_attempt_does_not_block_a_new_attempt(self):
        StudentAttempt.objects.create(
            student=self.student,
            exam_set=self.exam,
            is_complete=True,
        )

        StudentAttempt.objects.create(student=self.student, exam_set=self.exam)

    def test_only_one_answer_is_allowed_per_attempt_and_question(self):
        attempt = StudentAttempt.objects.create(student=self.student, exam_set=self.exam)
        StudentAnswer.objects.create(
            attempt=attempt,
            question=self.question,
            answer_text="answer",
        )

        with self.assertRaises(IntegrityError), transaction.atomic():
            StudentAnswer.objects.create(
                attempt=attempt,
                question=self.question,
                answer_text="duplicate",
            )

    def test_multiple_choice_configuration_is_validated(self):
        question = Question(
            section=self.section,
            order=2,
            question_type="mcq",
            prompt="Choose one.",
            options=["A"],
            correct_answer="B",
        )

        with self.assertRaises(ValidationError):
            question.full_clean()

    def test_manual_score_is_restricted_to_reviewed_response_types(self):
        attempt = StudentAttempt.objects.create(student=self.student, exam_set=self.exam)
        answer = StudentAnswer(
            attempt=attempt,
            question=self.question,
            answer_text="answer",
            manual_score=7.0,
        )

        with self.assertRaises(ValidationError):
            answer.full_clean()


class ExamNavigationTests(TestCase):
    def setUp(self):
        self.student = User.objects.create_user(username="navigator", password="test-pass-123")
        self.exam = ExamSet.objects.create(title="Ordered IELTS Mock", is_published=True)

        # Create these in reverse sequence to prove navigation does not rely on IDs.
        self.second_section = Section.objects.create(
            exam_set=self.exam,
            order=2,
            section_type="listening",
            time_limit_minutes=30,
        )
        self.first_section = Section.objects.create(
            exam_set=self.exam,
            order=1,
            section_type="reading",
            time_limit_minutes=60,
            passage_text="A short reading passage for interface testing.",
        )
        Question.objects.create(
            section=self.first_section,
            order=1,
            question_type="gap",
            prompt="First section question",
            correct_answer="answer",
        )
        Question.objects.create(
            section=self.second_section,
            order=1,
            question_type="gap",
            prompt="Second section question",
            correct_answer="answer",
        )
        self.client.force_login(self.student)

    def test_start_exam_opens_lowest_ordered_section(self):
        response = self.client.post(reverse("start_exam", args=[self.exam.id]))
        attempt = StudentAttempt.objects.get(student=self.student, exam_set=self.exam)

        self.assertRedirects(
            response,
            reverse("take_section", args=[attempt.id, self.first_section.id]),
            fetch_redirect_response=False,
        )

    def test_start_exam_rejects_get_requests(self):
        response = self.client.get(reverse("start_exam", args=[self.exam.id]))
        self.assertEqual(response.status_code, 405)

    def test_completed_test_can_be_retaken_without_replacing_the_old_attempt(self):
        completed_attempt = StudentAttempt.objects.create(
            student=self.student,
            exam_set=self.exam,
            is_complete=True,
            submitted_at=timezone.now(),
        )

        response = self.client.post(reverse("start_exam", args=[self.exam.id]))
        active_attempt = StudentAttempt.objects.get(
            student=self.student,
            exam_set=self.exam,
            is_complete=False,
        )

        self.assertNotEqual(active_attempt.id, completed_attempt.id)
        self.assertRedirects(
            response,
            reverse("take_section", args=[active_attempt.id, self.first_section.id]),
        )

    def test_retake_button_replaces_only_the_unfinished_attempt(self):
        active_attempt = StudentAttempt.objects.create(
            student=self.student,
            exam_set=self.exam,
            current_section=self.second_section,
            section_started_at=timezone.now(),
            section_deadline=timezone.now() + timedelta(minutes=30),
        )

        response = self.client.post(reverse("retake_exam", args=[active_attempt.id]))
        fresh_attempt = StudentAttempt.objects.get(
            student=self.student,
            exam_set=self.exam,
            is_complete=False,
        )

        self.assertFalse(StudentAttempt.objects.filter(id=active_attempt.id).exists())
        self.assertNotEqual(fresh_attempt.id, active_attempt.id)
        self.assertEqual(fresh_attempt.current_section, self.first_section)
        self.assertRedirects(
            response,
            reverse("take_section", args=[fresh_attempt.id, self.first_section.id]),
        )

    def test_reading_section_uses_fullscreen_workspace_and_question_navigator(self):
        self.client.post(reverse("start_exam", args=[self.exam.id]))
        attempt = StudentAttempt.objects.get(student=self.student, exam_set=self.exam)

        response = self.client.get(
            reverse("take_section", args=[attempt.id, self.first_section.id])
        )

        self.assertContains(response, "exam-session--reading")
        self.assertContains(response, 'id="fullscreen-toggle"')
        self.assertContains(response, 'class="exam-question-nav"')
        self.assertContains(response, 'data-target="question-')
        self.assertContains(response, 'id="exam-resizer"')
        self.assertContains(response, 'id="passage-selection-toolbar"')
        self.assertContains(response, 'id="passage-highlight"')
        self.assertContains(response, 'id="passage-add-note"')
        self.assertContains(response, 'id="passage-clear"')

    def test_submitting_section_opens_next_ordered_section(self):
        attempt = StudentAttempt.objects.create(student=self.student, exam_set=self.exam)
        attempt.current_section = self.first_section
        attempt.section_started_at = timezone.now()
        attempt.section_deadline = timezone.now() + timedelta(minutes=60)
        attempt.save()
        response = self.client.post(
            reverse("take_section", args=[attempt.id, self.first_section.id]),
            data={},
        )

        self.assertRedirects(
            response,
            reverse("take_section", args=[attempt.id, self.second_section.id]),
            fetch_redirect_response=False,
        )


class ExamCatalogueTests(TestCase):
    def setUp(self):
        self.student = User.objects.create_user(
            username="catalogue", password="test-pass-123"
        )
        self.exam = ExamSet.objects.create(
            title="Academic Practice Test 1", is_published=True
        )
        section = Section.objects.create(
            exam_set=self.exam,
            order=1,
            section_type="reading",
            time_limit_minutes=60,
        )
        Question.objects.create(
            section=section,
            order=1,
            question_type="gap",
            prompt="Complete the answer.",
            correct_answer="answer",
        )
        self.client.force_login(self.student)

    def test_catalogue_shows_real_test_metadata(self):
        response = self.client.get(reverse("exam_list"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Academic Practice Test 1")
        self.assertContains(response, "1</strong> section")
        self.assertContains(response, "1</strong> question")
        self.assertContains(response, "60</strong> min")
        self.assertNotContains(response, "рџ")

    def test_catalogue_status_filter_uses_the_students_attempt(self):
        StudentAttempt.objects.create(student=self.student, exam_set=self.exam)

        response = self.client.get(reverse("exam_list") + "?status=in_progress")

        self.assertContains(response, "Academic Practice Test 1")
        self.assertContains(response, "Continue test")
        self.assertEqual(response.context["counts"]["in_progress"], 1)

    def test_invalid_filter_falls_back_to_all(self):
        response = self.client.get(reverse("exam_list") + "?status=unknown")

        self.assertEqual(response.context["status_filter"], "all")

    def test_skill_filter_shows_only_the_selected_category(self):
        self.exam.category = "full"
        self.exam.save(update_fields=["category"])
        reading_exam = ExamSet.objects.create(
            title="Reading Only", category="reading", is_published=True
        )
        reading_section = Section.objects.create(
            exam_set=reading_exam,
            order=1,
            section_type="reading",
            time_limit_minutes=20,
        )
        Question.objects.create(
            section=reading_section,
            order=1,
            question_type="gap",
            prompt="Answer",
            correct_answer="yes",
        )

        response = self.client.get(reverse("exam_list") + "?skill=reading")

        self.assertContains(response, "Reading Only")
        self.assertNotContains(response, self.exam.title)
        self.assertEqual(response.context["skill_filter"], "reading")
        self.assertEqual(response.context["skill_counts"]["reading"], 1)

    def test_unpublished_tests_are_not_visible_or_startable(self):
        self.exam.is_published = False
        self.exam.save(update_fields=["is_published"])

        catalogue = self.client.get(reverse("exam_list"))
        start = self.client.post(reverse("start_exam", args=[self.exam.id]))

        self.assertNotContains(catalogue, self.exam.title)
        self.assertEqual(start.status_code, 404)


class ServerDeadlineTests(TestCase):
    def setUp(self):
        self.student = User.objects.create_user(username="timed", password="test-pass-123")
        self.exam = ExamSet.objects.create(title="Timed IELTS Mock", is_published=True)
        self.section = Section.objects.create(
            exam_set=self.exam,
            order=1,
            section_type="reading",
            time_limit_minutes=60,
            passage_text="The evidence sentence is in this reading passage.",
        )
        self.question = Question.objects.create(
            section=self.section,
            order=1,
            question_type="gap",
            prompt="Complete this answer.",
            correct_answer="valid",
            explanation="The passage gives the exact word used in the answer.",
            passage_reference="The evidence sentence is in this reading passage.",
        )
        self.client.force_login(self.student)

    def start_attempt(self):
        self.client.post(reverse("start_exam", args=[self.exam.id]))
        return StudentAttempt.objects.get(student=self.student, exam_set=self.exam)

    def test_refresh_does_not_reset_server_deadline(self):
        attempt = self.start_attempt()
        original_deadline = attempt.section_deadline

        self.client.get(reverse("take_section", args=[attempt.id, self.section.id]))
        attempt.refresh_from_db()

        self.assertEqual(attempt.section_deadline, original_deadline)

    def test_late_submission_does_not_accept_new_answer_text(self):
        attempt = self.start_attempt()
        late_time = attempt.section_deadline + timedelta(seconds=SUBMISSION_GRACE_SECONDS + 1)

        with patch("exams.views.timezone.now", return_value=late_time):
            self.client.post(
                reverse("take_section", args=[attempt.id, self.section.id]),
                data={f"q{self.question.id}": "valid"},
            )

        answer = StudentAnswer.objects.get(attempt=attempt, question=self.question)
        self.assertEqual(answer.answer_text, "")
        self.assertFalse(answer.is_correct)

    def test_user_cannot_skip_to_a_later_section(self):
        later_section = Section.objects.create(
            exam_set=self.exam,
            order=2,
            section_type="listening",
            time_limit_minutes=30,
        )
        Question.objects.create(
            section=later_section,
            order=1,
            question_type="gap",
            prompt="Later question",
            correct_answer="answer",
        )
        attempt = self.start_attempt()

        response = self.client.get(
            reverse("take_section", args=[attempt.id, later_section.id])
        )

        self.assertRedirects(
            response,
            reverse("take_section", args=[attempt.id, self.section.id]),
            fetch_redirect_response=False,
        )

    def test_completion_awards_xp_and_updates_streak_once(self):
        attempt = self.start_attempt()
        response = self.client.post(
            reverse("take_section", args=[attempt.id, self.section.id]),
            data={f"q{self.question.id}": "valid"},
        )

        self.student.studentprofile.refresh_from_db()
        self.assertRedirects(
            response,
            reverse("results", args=[attempt.id]),
            fetch_redirect_response=False,
        )
        self.assertEqual(self.student.studentprofile.xp, 60)
        self.assertEqual(self.student.studentprofile.streak, 1)
        self.assertEqual(self.student.studentprofile.last_activity_date, timezone.localdate())

    def test_results_review_shows_correct_answer_explanation_and_passage_link(self):
        attempt = self.start_attempt()
        self.client.post(
            reverse("take_section", args=[attempt.id, self.section.id]),
            data={f"q{self.question.id}": "wrong"},
        )

        response = self.client.get(reverse("results", args=[attempt.id]))

        self.assertContains(response, "Correct answer")
        self.assertNotContains(response, self.question.explanation)
        self.assertContains(response, self.question.passage_reference)
        self.assertContains(response, 'class="reader-show-evidence"')
        self.assertContains(response, f'data-passage-section="{self.section.id}"')


class DiagnosticSeedTests(TestCase):
    def test_seed_command_creates_one_ready_published_diagnostic(self):
        call_command("seed_diagnostic_exam", verbosity=0)
        call_command("seed_diagnostic_exam", verbosity=0)

        exam = ExamSet.objects.get(title="IELTS Skills Diagnostic")
        self.assertTrue(exam.is_published)
        self.assertTrue(exam.is_ready)
        self.assertEqual(exam.sections.count(), 3)
        self.assertEqual(Question.objects.filter(section__exam_set=exam).count(), 5)


class PracticeLibrarySeedTests(TestCase):
    def setUp(self):
        self.media_directory = TemporaryDirectory()
        self.media_override = override_settings(MEDIA_ROOT=self.media_directory.name)
        self.media_override.enable()

    def tearDown(self):
        self.media_override.disable()
        self.media_directory.cleanup()
        super().tearDown()

    def test_seed_library_creates_all_five_separate_categories(self):
        call_command("seed_practice_library", verbosity=0)
        call_command("seed_practice_library", verbosity=0)

        published = ExamSet.objects.filter(is_published=True)
        self.assertEqual(published.count(), 5)
        self.assertEqual(
            set(published.values_list("category", flat=True)),
            {"reading", "listening", "writing", "speaking", "full"},
        )
        full_mock = published.get(category="full")
        self.assertEqual(
            list(full_mock.sections.values_list("section_type", flat=True)),
            ["listening", "reading", "writing", "speaking"],
        )
        listening = published.get(category="listening").sections.get()
        self.assertTrue(bool(listening.audio_file))


class CompleteDiagnosticJourneyTests(TestCase):
    def setUp(self):
        self.media_directory = TemporaryDirectory()
        self.media_override = override_settings(MEDIA_ROOT=self.media_directory.name)
        self.media_override.enable()
        call_command("seed_diagnostic_exam", verbosity=0)
        self.exam = ExamSet.objects.get(title="IELTS Skills Diagnostic")
        self.student = User.objects.create_user(
            username="journey", password="test-pass-123"
        )
        self.client.force_login(self.student)

    def tearDown(self):
        self.media_override.disable()
        self.media_directory.cleanup()
        super().tearDown()

    def test_student_can_complete_the_published_diagnostic(self):
        detail = self.client.get(reverse("exam_detail", args=[self.exam.id]))
        self.assertContains(detail, "Begin test")

        start = self.client.post(reverse("start_exam", args=[self.exam.id]))
        attempt = StudentAttempt.objects.get(student=self.student, exam_set=self.exam)
        reading, writing, speaking = list(self.exam.sections.all())
        self.assertRedirects(
            start,
            reverse("take_section", args=[attempt.id, reading.id]),
            fetch_redirect_response=False,
        )

        reading_answers = {
            f"q{question.id}": question.correct_answer
            for question in reading.questions.all()
        }
        reading_submit = self.client.post(
            reverse("take_section", args=[attempt.id, reading.id]),
            data=reading_answers,
        )
        self.assertRedirects(
            reading_submit,
            reverse("take_section", args=[attempt.id, writing.id]),
            fetch_redirect_response=False,
        )

        writing_question = writing.questions.get()
        self.client.post(
            reverse("take_section", args=[attempt.id, writing.id]),
            data={f"q{writing_question.id}": "A structured diagnostic essay response."},
        )
        speaking_question = speaking.questions.get()
        finish = self.client.post(
            reverse("take_section", args=[attempt.id, speaking.id]),
            data={f"q{speaking_question.id}": "Structured notes for a spoken response."},
            follow=True,
        )

        attempt.refresh_from_db()
        self.student.studentprofile.refresh_from_db()
        self.assertTrue(attempt.is_complete)
        self.assertEqual(attempt.answers.count(), 5)
        self.assertEqual(attempt.answers.filter(is_correct=True).count(), 3)
        self.assertEqual(self.student.studentprofile.xp, 100)
        self.assertContains(finish, "3 / 3")
        self.assertContains(finish, "Awaiting review")
        self.assertContains(finish, "Pending instructor review", count=2)

    def test_speaking_audio_upload_is_stored_and_invalid_types_are_rejected(self):
        self.client.post(reverse("start_exam", args=[self.exam.id]))
        attempt = StudentAttempt.objects.get(student=self.student, exam_set=self.exam)
        reading, writing, speaking = list(self.exam.sections.all())
        reading_answers = {
            f"q{question.id}": question.correct_answer
            for question in reading.questions.all()
        }
        self.client.post(
            reverse("take_section", args=[attempt.id, reading.id]),
            data=reading_answers,
        )
        writing_question = writing.questions.get()
        self.client.post(
            reverse("take_section", args=[attempt.id, writing.id]),
            data={f"q{writing_question.id}": "Essay response"},
        )
        speaking_question = speaking.questions.get()
        invalid = SimpleUploadedFile("response.exe", b"not audio")
        rejected = self.client.post(
            reverse("take_section", args=[attempt.id, speaking.id]),
            data={f"q{speaking_question.id}_audio": invalid},
            follow=True,
        )
        self.assertContains(rejected, "File extension")
        attempt.refresh_from_db()
        self.assertFalse(attempt.is_complete)

        valid = SimpleUploadedFile("response.webm", b"small audio placeholder", "audio/webm")
        self.client.post(
            reverse("take_section", args=[attempt.id, speaking.id]),
            data={f"q{speaking_question.id}_audio": valid},
        )
        answer = StudentAnswer.objects.get(attempt=attempt, question=speaking_question)
        self.assertTrue(answer.audio_response.name.endswith(".webm"))
