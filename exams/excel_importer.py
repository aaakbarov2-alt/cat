from io import BytesIO
import re

from django.core.exceptions import ValidationError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


TEST_FIELDS = {"title", "description", "category"}
SECTION_HEADERS = ["section_order", "section_type", "time_limit_minutes", "passage_text"]
QUESTION_HEADERS = [
    "section_order", "question_order", "question_type", "prompt", "options",
    "correct_answer", "explanation", "passage_reference", "notes_for_admin",
]
QUESTION_HEADERS_GROUPED = QUESTION_HEADERS + ["group_key"]
GROUP_HEADERS = ["section_order", "group_key", "group_order", "layout_type", "title", "instructions", "layout_html"]
SKILLS = {"reading", "listening", "writing", "speaking", "full"}
SECTION_TYPES = {"reading", "listening", "writing", "speaking"}
QUESTION_TYPES = {"mcq", "gap", "matching", "essay", "speaking"}


class ExcelImportError(ValidationError):
    pass


def _text(value):
    return "" if value is None else str(value).strip()


def _integer(value, label, row, minimum=1, maximum=240):
    try:
        if isinstance(value, float) and value.is_integer():
            number = int(value)
        elif isinstance(value, str):
            match = re.fullmatch(r"\s*(\d+)(?:\.0+)?(?:\s*(?:min|mins|minute|minutes))?\s*", value, re.IGNORECASE)
            if not match:
                raise ValueError
            number = int(match.group(1))
        else:
            number = int(value)
    except (TypeError, ValueError):
        raise ExcelImportError(f"{label}, row {row}: enter a whole number.")
    if not minimum <= number <= maximum:
        raise ExcelImportError(f"{label}, row {row}: use a number from {minimum} to {maximum}.")
    return number


def _headers(sheet, expected):
    actual = [_text(cell.value).lower() for cell in sheet[1]][:len(expected)]
    if actual != expected:
        raise ExcelImportError(
            f"Sheet '{sheet.title}' has changed column headers. Download a fresh template and keep its first row unchanged."
        )


def parse_excel_test(file_object, publish=False):
    try:
        workbook = load_workbook(file_object, read_only=True, data_only=True)
    except Exception as error:
        raise ExcelImportError(f"The workbook could not be opened: {error}")
    required = {"Test", "Sections", "Questions"}
    missing = required.difference(workbook.sheetnames)
    if missing:
        raise ExcelImportError("Missing worksheet(s): " + ", ".join(sorted(missing)))

    test_sheet = workbook["Test"]
    metadata = {}
    for row in test_sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        key = _text(row[0]).lower()
        if key:
            metadata[key] = _text(row[1])
    if not metadata.get("title"):
        raise ExcelImportError("Test sheet: title is required.")
    category = metadata.get("category", "full").lower()
    if category not in SKILLS:
        raise ExcelImportError("Test sheet: category must be reading, listening, writing, speaking, or full.")

    section_sheet = workbook["Sections"]
    if section_sheet.max_row is not None and section_sheet.max_row > 51:
        raise ExcelImportError("Sections sheet: maximum 50 section rows are allowed.")
    _headers(section_sheet, SECTION_HEADERS)
    sections = {}
    for row_number, row in enumerate(section_sheet.iter_rows(min_row=2, max_col=4, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        order = _integer(row[0], "Sections sheet section_order", row_number, maximum=50)
        section_type = _text(row[1]).lower()
        if section_type not in SECTION_TYPES:
            raise ExcelImportError(f"Sections sheet, row {row_number}: invalid section_type '{section_type}'.")
        if order in sections:
            raise ExcelImportError(f"Sections sheet: section_order {order} appears more than once.")
        sections[order] = {
            "order": order,
            "section_type": section_type,
            "time_limit_minutes": (
                _integer(row[2], "Sections sheet time_limit_minutes", row_number)
                if _text(row[2])
                else {"reading": 60, "listening": 30, "writing": 60, "speaking": 15}[section_type]
            ),
            "passage_text": _text(row[3]),
            "questions": [],
        }
    if not sections:
        raise ExcelImportError("Sections sheet: add at least one section.")

    if "Groups" in workbook.sheetnames:
        group_sheet = workbook["Groups"]
        _headers(group_sheet, GROUP_HEADERS)
        for row_number, row in enumerate(group_sheet.iter_rows(min_row=2, max_col=7, values_only=True), start=2):
            if not any(value not in (None, "") for value in row):
                continue
            section_order = _integer(row[0], "Groups sheet section_order", row_number, maximum=50)
            if section_order not in sections:
                raise ExcelImportError(f"Groups sheet, row {row_number}: section_order {section_order} is not on the Sections sheet.")
            key = _text(row[1]).lower()
            if not re.fullmatch(r"[a-z0-9_-]+", key):
                raise ExcelImportError(f"Groups sheet, row {row_number}: group_key must use letters, numbers, hyphens, or underscores.")
            layout_type = _text(row[3]).lower() or "notes"
            if layout_type not in {"notes", "table", "flow"}:
                raise ExcelImportError(f"Groups sheet, row {row_number}: layout_type must be notes, table, or flow.")
            if any(group["key"] == key for group in sections[section_order].setdefault("groups", [])):
                raise ExcelImportError(f"Groups sheet, row {row_number}: group_key '{key}' is duplicated in section {section_order}.")
            layout_html = _text(row[6])
            if not layout_html:
                raise ExcelImportError(f"Groups sheet, row {row_number}: layout_html is required.")
            sections[section_order].setdefault("groups", []).append({
                "key": key, "order": _integer(row[2], "Groups sheet group_order", row_number, maximum=200),
                "layout_type": layout_type, "title": _text(row[4]), "instructions": _text(row[5]), "layout_html": layout_html,
            })

    question_sheet = workbook["Questions"]
    if question_sheet.max_row is not None and question_sheet.max_row > 251:
        raise ExcelImportError("Questions sheet: maximum 250 question rows are allowed.")
    actual_headers = [_text(cell.value).lower() for cell in question_sheet[1]]
    if actual_headers[:len(QUESTION_HEADERS_GROUPED)] == QUESTION_HEADERS_GROUPED:
        question_columns = 10
    elif actual_headers[:len(QUESTION_HEADERS)] == QUESTION_HEADERS:
        question_columns = 9
    else:
        raise ExcelImportError("Sheet 'Questions' has changed column headers. Download a fresh template and keep its first row unchanged.")
    seen = set()
    for row_number, row in enumerate(question_sheet.iter_rows(min_row=2, max_col=question_columns, values_only=True), start=2):
        if not any(value not in (None, "") for value in row):
            continue
        section_order = _integer(row[0], "Questions sheet section_order", row_number, maximum=50)
        if section_order not in sections:
            raise ExcelImportError(f"Questions sheet, row {row_number}: section_order {section_order} is not on the Sections sheet.")
        question_order = _integer(row[1], "Questions sheet question_order", row_number, maximum=200)
        key = (section_order, question_order)
        if key in seen:
            raise ExcelImportError(f"Questions sheet: question_order {question_order} is duplicated in section {section_order}.")
        seen.add(key)
        question_type = _text(row[2]).lower()
        if question_type not in QUESTION_TYPES:
            raise ExcelImportError(f"Questions sheet, row {row_number}: invalid question_type '{question_type}'.")
        prompt = _text(row[3])
        if not prompt:
            raise ExcelImportError(f"Questions sheet, row {row_number}: prompt is required.")
        options = [item.strip() for item in _text(row[4]).split("|") if item.strip()]
        answer = _text(row[5])
        if question_type in {"mcq", "matching"}:
            if len(options) < 2:
                raise ExcelImportError(f"Questions sheet, row {row_number}: {question_type} requires at least two | separated options.")
            if answer not in options:
                raise ExcelImportError(f"Questions sheet, row {row_number}: correct_answer must exactly match one option.")
        if question_type == "gap" and not answer:
            raise ExcelImportError(f"Questions sheet, row {row_number}: gap questions require a correct_answer.")
        group_key = _text(row[9]).lower() if question_columns == 10 else ""
        if group_key:
            groups = sections[section_order].get("groups", [])
            group = next((item for item in groups if item["key"] == group_key), None)
            if group is None:
                raise ExcelImportError(f"Questions sheet, row {row_number}: group_key '{group_key}' is not defined on the Groups sheet.")
            if question_type not in {"gap", "matching"}:
                raise ExcelImportError(f"Questions sheet, row {row_number}: grouped questions must use gap or matching.")
            if f"[[{question_order}]]" not in group["layout_html"]:
                raise ExcelImportError(f"Questions sheet, row {row_number}: group '{group_key}' layout_html does not contain [[{question_order}]].")
        sections[section_order]["questions"].append({
            "order": question_order,
            "question_type": question_type,
            "prompt": prompt,
            "options": options or None,
            "correct_answer": answer or None,
            "explanation": _text(row[6]),
            "passage_reference": _text(row[7]),
            "group_key": group_key,
        })
    empty_sections = [str(order) for order, section in sections.items() if not section["questions"]]
    if empty_sections:
        raise ExcelImportError("Every section needs at least one question. Empty section_order: " + ", ".join(empty_sections))
    for section in sections.values():
        for group in section.get("groups", []):
            grouped_orders = {question["order"] for question in section["questions"] if question.get("group_key") == group["key"]}
            placeholder_orders = [int(value) for value in re.findall(r"\[\[(\d+)\]\]", group["layout_html"])]
            if len(placeholder_orders) != len(set(placeholder_orders)):
                raise ExcelImportError(f"Groups sheet: group '{group['key']}' contains a duplicate question placeholder.")
            if set(placeholder_orders) != grouped_orders:
                raise ExcelImportError(f"Groups sheet: group '{group['key']}' placeholders must exactly match its Questions rows.")
        section["questions"].sort(key=lambda item: item["order"])

    return {
        "title": metadata["title"],
        "description": metadata.get("description", ""),
        "category": category,
        "delivery_mode": "native",
        "publish": bool(publish),
        "sections": [sections[key] for key in sorted(sections)],
    }


def build_excel_template():
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "Instructions"
    test = workbook.create_sheet("Test")
    sections = workbook.create_sheet("Sections")
    questions = workbook.create_sheet("Questions")
    groups = workbook.create_sheet("Groups")
    navy, blue, pale = "10213E", "1463E9", "EAF2FF"

    instructions.append(["IELTS Mock — Excel Import Template"])
    instructions.append(["Keep sheet names and column headers unchanged."])
    instructions.append(["1", "Complete Test, Sections, and Questions. Use Groups only for inline notes, tables, forms, or flow charts."])
    instructions.append(["2", "Use | between MCQ or matching options."])
    instructions.append(["3", "Upload in Admin → Import Excel tests, review, then save."])
    instructions.column_dimensions["A"].width = 16
    instructions.column_dimensions["B"].width = 78

    test.append(["field", "value"])
    test.append(["title", "Academic Reading Practice 1"])
    test.append(["description", "A complete IELTS practice test imported from Excel."])
    test.append(["category", "reading"])
    test.column_dimensions["A"].width = 22
    test.column_dimensions["B"].width = 68

    sections.append(SECTION_HEADERS)
    sections.append([1, "reading", 20, "Paste the complete reading passage here."])
    sections.column_dimensions["A"].width = 16
    sections.column_dimensions["B"].width = 20
    sections.column_dimensions["C"].width = 22
    sections.column_dimensions["D"].width = 75

    groups.append(GROUP_HEADERS)
    groups.append([1, "notes_1", 1, "notes", "Urban farming in Paris", "Complete the notes below. Choose NO MORE THAN TWO WORDS AND/OR A NUMBER.", "<h3>Urban farming in Paris</h3><h4>Farm layout and production</h4><ul><li>Vertical tubes grow strawberries, [[1]] and herbs.</li><li>The daily harvest may reach [[2]] in weight.</li></ul>"])
    for index, width in enumerate([15, 18, 15, 18, 32, 65, 95], start=1):
        groups.column_dimensions[chr(64 + index)].width = width
    questions.append(QUESTION_HEADERS_GROUPED)
    questions.append([1, 1, "gap", "Vertical tubes grow strawberries, _____ and herbs.", "", "lettuces", "Explain why Option B is correct.", "From identical vertical tubes nearby burst row upon row of lettuces.", "", "notes_1"])
    questions.append([1, 2, "gap", "The daily harvest may reach _____ in weight.", "", "1,000 kg", "", "Staff will harvest up to 1,000 kg every day.", "", "notes_1"])
    widths = [15, 15, 18, 48, 44, 28, 42, 46, 28, 20]
    for index, width in enumerate(widths, start=1):
        questions.column_dimensions[chr(64 + index)].width = width

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=navy)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    category_validation = DataValidation(type="list", formula1='"reading,listening,writing,speaking,full"')
    test.add_data_validation(category_validation)
    category_validation.add(test["B4"])
    section_validation = DataValidation(type="list", formula1='"reading,listening,writing,speaking"')
    sections.add_data_validation(section_validation)
    section_validation.add("B2:B50")
    question_validation = DataValidation(type="list", formula1='"mcq,gap,matching,essay,speaking"')
    questions.add_data_validation(question_validation)
    question_validation.add("C2:C250")
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
